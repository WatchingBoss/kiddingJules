from openpyxl import load_workbook
import sys

wb = load_workbook('test.xlsx')
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    for row in ws.iter_rows(values_only=True):
        print(row)
